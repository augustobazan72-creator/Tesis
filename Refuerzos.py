import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import gc
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
import copy
import pandapower as pp
from datetime import timedelta
from haversine import haversine, Unit
from Analisis_estadistico import (analisis_caso_base, percentiles_cb, constantes, ALTO, ANCHO, DPI,
                                aplicar_tema_light, analisis_contingencias, indice_cond_n)
from Red_pandapower import lista_doble_terna, reporte_red, trafos_gen
from Motor_DC import (Configuracion_Simulacion, Configuracion_Simulacion_Contingencias, caso_base_completo,
                    contingencias_refuerzos)
from Rutas import creacion_carpetas_refuerzos
from Menus import prints_refuerzos_usuario
from Configuracion_inicial import input_log
from Lector_excels import convertir_a_lista
from Analisis_economico import calculo_inversion
from Resultados import resultados_alternativas_propuestas

# --- Configuracion logging ---
logger = logging.getLogger(__name__)

# --- ACTIVACION DEL REFUERZO PROPUESTO ---
def activar_ref_propuesto (id_ref, tipo, nombre, net_copy):
    if tipo=='line':
        parametros_ref = net_copy.line.loc[id_ref].copy() # copiamos los parametros de la linea original
        datos_dict = parametros_ref.to_dict() # lo convertimos a diccionario
        pp.create_line_from_parameters(net_copy,
            name=nombre,
            from_bus=datos_dict['from_bus'],
            to_bus=datos_dict['to_bus'],
            r_ohm_per_km = datos_dict['r_ohm_per_km'],
            x_ohm_per_km = datos_dict['x_ohm_per_km'],
            c_nf_per_km = datos_dict['c_nf_per_km'],
            max_i_ka = datos_dict['max_i_ka'],
            in_service = datos_dict['in_service'], 
            g_us_per_km = datos_dict['g_us_per_km'],
            type = datos_dict['type'],
            df=datos_dict['df'],
            length_km = datos_dict['length_km'],
            parallel = datos_dict['parallel'],
            max_loading_percent=datos_dict['max_loading_percent'])
        logger.info(f'Se conecto la propuesta de refuerzo a la red base: {nombre} - tipo: {tipo}')
    else: 
        parametros_ref = net_copy.trafo.loc[id_ref].copy() # copiamos los parametros del trafo original
        datos_dict = parametros_ref.to_dict() # lo llevamos a diccioario
        pp.create_transformer_from_parameters(net_copy,
            name = nombre,
            hv_bus = datos_dict['hv_bus'],
            lv_bus = datos_dict['lv_bus'],
            sn_mva = datos_dict['sn_mva'],
            vn_hv_kv = datos_dict['vn_hv_kv'],
            vn_lv_kv = datos_dict['vn_lv_kv'],
            vk_percent = datos_dict['vk_percent'],
            vkr_percent = datos_dict['vkr_percent'],
            in_service = datos_dict['in_service'],
            pfe_kw = datos_dict['pfe_kw'],
            i0_percent = datos_dict['i0_percent'],
            shift_degree = datos_dict['shift_degree'],
            max_loading_percent = datos_dict['max_loading_percent'],
            parallel = datos_dict['parallel'])
        logger.info(f'Se conecto la propuesta de refuerzo a la red base: {nombre} - tipo: {tipo}')

# --- IDENTIFICACION DE LOS REFUERZOS A PROPONER ---

def elementos_criticos (net, analisis_componentes, ranking_contingencias_cb):
    print(f'{'-'*80}')
    print('IDENTIFICANDO ELEMENTOS CRITICOS')
    print(f'{'-'*80}')
    # PREPARANDO LISTAS
    elementos_condicion_n = analisis_componentes['Nombre_Componente'].tolist()[:19]
    if not elementos_condicion_n:
        e_1 = 'La columna de [Nombre_Componente] no retorna valores.'
        logger.error(e_1)
        e_2 = 'Revisar que el df de analisis de componentes tenga datos'
        raise ValueError(e_2)
    df_cont = ranking_contingencias_cb.query('Ind_Sev >= 1')
    elementos_contingencias = df_cont['Contingencia'].tolist()[:19]
    if not elementos_contingencias:
        e_1 = 'La columna de [Contingencia] no retorna valores (Porque no hay datos o No se tienen contingencias con i_sev >= 1).'
        logger.error(e_1)
        e_2 = 'Revisar que el df de ranking de contingencias.'
        raise ValueError(e_2)
    elementos_criticos = set(elementos_condicion_n + elementos_contingencias)
    lineas_2t = lista_doble_terna(net)
    # PROCESAMIENTO DE LISTAS
    elementos_criticos_filtrados = elementos_criticos.difference(lineas_2t)
    elementos_criticos_filtrados = list(elementos_criticos_filtrados)
    if len(elementos_criticos_filtrados) > 30:
        elementos_criticos_filtrados = elementos_criticos_filtrados[:29]
    logger.info(f'El analisis de refuerzos se hara para {len(elementos_criticos_filtrados)} elementos criticos,\n{elementos_criticos}')
    lineas, trafos =  [], []
    for elemento in elementos_criticos_filtrados:
        if str(elemento[0:3]) != str(elemento[6:9]):
            lineas.append(elemento)
        else: 
            trafos.append(elemento)
    """
    No se hace ningun filtrado a los trafos porque estos ya se filtraron antes de ejecutar el analisis
    tanto estadistico como de contingencias
    """
    datos_refuerzos = []
    for refuerzo in elementos_criticos_filtrados:
        nombre = str(refuerzo) + '_r'
        idx_line = net.line[net.line['name'] == refuerzo].index
        if not idx_line.empty:
            datos_refuerzos.append((idx_line[0], 'line', nombre))
            continue
        idx_trafo = net.trafo[net.trafo['name'] == refuerzo].index
        if not idx_trafo.empty:
            datos_refuerzos.append((idx_trafo[0], 'trafo', nombre))
            continue
    print(f'{'-'*80}')
    return datos_refuerzos

# --- UNIMOS DF CARGABILIDADES CON LA DURACION POR BLOQUE ---
def unir_duracion(df_duraci: pd.DataFrame, df:pd.DataFrame):
    df_duracion = (df_duraci.drop_duplicates(subset=['Bloque']).drop(columns=['Etapa','Serie']).copy()) # Hrs/bloq
    df_unido = pd.merge(df, df_duracion, on='Bloque', how='left')
    return df_unido

# --- IDENTIFICACION DE ELEMENTOS SENSIBLES A LA PROPUESTA ---
def elementos_monitoreo(df_cargabilidades_base: pd.DataFrame, df_cargabilidades_ref: pd.DataFrame,
                        n_elementos: int, nombre:str):
    print(f'{'='*80}')
    print(f'EVALUACION DEL REFUERZO EN CONDICION N.')
    print(f'{'='*80}')
    df_cargab_union = pd.merge(df_cargabilidades_base, df_cargabilidades_ref, how='inner',
                            on=['Etapa', 'Serie', 'Bloque','Componente'],
                            suffixes= ('_base', '_ref'))
    df_cargab_union = df_cargab_union[['Etapa', 'Serie', 'Bloque', 'Componente', 'loading_percent_base', 
                                    'loading_percent_ref']].copy()
    df_cargab_union ['delta_cargab'] = (df_cargab_union['loading_percent_base'] - df_cargab_union['loading_percent_ref']).abs()
    df_maximos = df_cargab_union.groupby('Componente')['delta_cargab'].max().reset_index()
    df_maximos.sort_values(by='delta_cargab', ascending=False, inplace=True)
    df_maximos = df_maximos.head(n_elementos)
    elementos_sensibles = df_maximos['Componente'].to_list()
    logger.info(f'Los {n_elementos} elementos de monitoreo para {nombre} son:\n{elementos_sensibles}')
    return elementos_sensibles

# --- EVALUACION DE LA CONDICION N CON REFUERZO VS SIN REFUERZO ---
def comparativa_cn(df_cargabilidades_base: pd.DataFrame, df_cargabilidades_ref: pd.DataFrame,
                    datos_estudio : dict, nombre: str, horas_serie: int, lista_monitoreados: list, 
                    rta_cb_r: str|Path, estudio:int):
    # EXTRAEMOS LOS PERCENTILES POR ELEMENTO 
    evaluacion_cb = []
    for elemento in lista_monitoreados:
        p1_cr, p2_cr = percentiles_cb(df_cargabilidades_ref, elemento, datos_estudio['numero_series'], horas_serie)
        p1_sr, p2_sr = percentiles_cb(df_cargabilidades_base, elemento, datos_estudio['numero_series'], horas_serie)
        fila = [nombre, elemento, p1_sr, p2_sr, p1_cr, p2_cr]
        evaluacion_cb.append(fila)
    # ARMAMOS EL DATAFRAME
    if estudio == 1:
        encabezados = ['Refuerzo', 'Monitoreo', 'P_1%(SR)', 'P_5%(SR)', 'P_1%(CR)','P_5%(CR)']
    else :
        encabezados = ['Alternativa', 'Monitoreo', 'P_1%(SR)', 'P_5%(SR)', 'P_1%(CR)','P_5%(CR)']
    df_condicion_n = pd.DataFrame(evaluacion_cb, columns=encabezados)
    df_condicion_n['Alivio(P_5%)'] = df_condicion_n['P_5%(SR)'] - df_condicion_n['P_5%(CR)']
    logger.info('Se genero correctamente el reporte tabular de la evaluacion de la propuesta.')
        # PARA LA CONSTRUCCION DEL DF RESULTADOS
    if estudio == 1:
        df_condicion_n.to_csv((Path(rta_cb_r)/'Evaluacion_tecnica_condicon_n.csv'), index=False)
        fila_res = df_condicion_n.loc[df_condicion_n['Monitoreo'] == nombre[:-2]].iloc[0]
        return df_condicion_n, fila_res
    else:
        df_condicion_n.to_csv((Path(rta_cb_r)/f'Evaluacion_tecnica_{nombre}_(Cond_n).csv'), index=False)
        return df_condicion_n

# --- Comparativa propuesta reb base vs con refuerzo ----
def grafica_comparativa_cb(df: pd.DataFrame, nombre: str, ruta_caso_base: str|Path):
    nombre_graf = f'Comparativa_(cond_n)_{nombre}'
    ruta = Path(ruta_caso_base)
    componentes = df['Monitoreo'].astype(str).tolist()
    p1_sr = df['P_1%(SR)'].astype(float).to_numpy()
    p1_cr = df['P_1%(CR)'].astype(float).to_numpy()
    p5_sr = df['P_5%(SR)'].astype(float).to_numpy()
    p5_cr = df['P_5%(CR)'].astype(float).to_numpy()
    x = np.arange(len(componentes))
    ancho_barra = 0.2
    fig, ax = plt.subplots(figsize=(ANCHO + 2.2, ALTO))
    _ = ax.bar(x - 1.5 * ancho_barra, p1_sr, ancho_barra, label='P_1%(SR)')
    _ = ax.bar(x - 0.5 * ancho_barra, p1_cr, ancho_barra, label='P_1%(CR)')
    _ = ax.bar(x + 0.5 * ancho_barra, p5_sr, ancho_barra, label='P_5%(SR)')
    _ = ax.bar(x + 1.5 * ancho_barra, p5_cr, ancho_barra, label='P_5%(CR)')
    ax.set_ylim(0, p1_sr.max() + 10)
    ax.set_xticks(x)
    ax.set_xticklabels(componentes, rotation=45, ha='right')
    ax.set_xlabel('Componentes de monitoreo.')
    ax.set_ylabel('Cargabilidad [%]')
    ax.legend()
    aplicar_tema_light(
        ax,
        titulo=f'Impacto de la propuesta {nombre}.',
        subtitulo='Comparativa condicion "n".',
        xlabel='Componentes de monitoreo.',
        ylabel='Cargabilidad [%]'
    )
    ax.tick_params(axis = 'both', labelsize = 7)
    plt.tight_layout()
    plt.savefig(ruta / f"{nombre_graf}.png", dpi=DPI)
    plt.close()
    logger.info('Se genero la grafica comparativa parala condicion n')
    print(f'{'='*80}')

def comparativa_ctg(ranking_contingencias_cb: pd.DataFrame, indice_severidad_ref: pd.DataFrame, nombre : str,
                ruta_cont : str | Path):
    print(f'{'='*80}')
    print('EVALUACION DEL REFUERZO EN CONDICION N-1 (CONTINGENCIAS).')
    print(f'{'='*80}')
    comparativa_is = pd.DataFrame()
    comparativa_is['Contingencia'] = ranking_contingencias_cb['Contingencia'].values
    comparativa_is['is_pre_ref'] = ranking_contingencias_cb['Ind_Sev'].values
    comparativa_is['pip_mx_pre_ref'] = ranking_contingencias_cb['PIp_max'].values
    comparativa_is = pd.merge(comparativa_is, indice_severidad_ref, how='inner', on='Contingencia')
    comparativa_is = comparativa_is.rename(columns={'Ind_Sev' : 'is_post_ref', 'PIp_max' : 'pip_mx_post_ref'})
    comparativa_is = comparativa_is [['Contingencia', 'is_pre_ref', 'pip_mx_pre_ref', 'is_post_ref', 
                                    'pip_mx_post_ref']].copy()
    comparativa_is['beneficio_tecnico_is'] = comparativa_is['is_pre_ref'] - comparativa_is['is_post_ref']
    # filtrar los datos
    nombre_arch = f'Evaluacion_isev_{nombre}.csv'
    logger.info(f'Se genero el reporte comparativo de indices de severidad tabular.')
    comparativa_is.to_csv((Path(ruta_cont)/nombre_arch) , index=False)
    # grafica
    fig, ax = plt.subplots(figsize = (ANCHO, ALTO))
    ax.plot(comparativa_is['Contingencia'],comparativa_is['is_pre_ref'], color = 'blue', linestyle = '--', 
            marker = '.', label='Pre-Refuerzo', zorder = 1)
    ax.plot(comparativa_is['Contingencia'],comparativa_is['is_post_ref'], color = 'red', linestyle = '-',
            marker = '.', label='Post-Refuerzo', zorder = 2)
    ax.legend()
    ax.grid(True, linestyle=':', alpha=0.6, zorder = 0)
    ax.set_ylabel('Indice de Seguridad (IS)')
    ax.set_xlabel('Contingencias Analizadas')
    ax.set_title(f'Comparativa de IS para el ref: {nombre}')
    ax.tick_params(axis = 'both', labelsize = 7)
    x = np.arange(len(comparativa_is['Contingencia'].tolist()))
    ax.set_xticks(x)
    ax.set_xticklabels(comparativa_is['Contingencia'], rotation=45, ha='right')
    fig.tight_layout()
    plt.savefig(f"{ruta_cont}/Grafica_comparativa_indice_severidad_{nombre}.png", dpi=DPI)
    plt.close()
    logger.info(f'Se genero la grafica comparativa de indices de severidad para {nombre}.')
    print(f'{'='*80}')

# --- FUNCION PRINCIPAL DE LA PROPUESTA DE REFUERZOS --- 
def analisis_ref_popuestos(net, analisis_componentes: pd.DataFrame, ranking_contingencias_cb: pd.DataFrame, ruta_refuerzos : Path|str,
                    df_mline: pd.DataFrame, df_mtrafo: pd.DataFrame, df_demanda: pd.DataFrame, df_desp_TH: pd.DataFrame,
                    df_desp_ren: pd.DataFrame, Slacks: pd.Series, datos_estudio: pd.DataFrame, df_fechas: pd.DataFrame, 
                    df_duraci: pd.DataFrame, config_red, exponente_n_pip, nucleos, df_cargabilidades_base, n_elementos,
                    trafos_limpios):
    
    print(f'\n{'='*80}')
    print(f'ANALISIS DE PROPUESTA DE REFUERZOS.')
    print(f'{'='*80}')

    # PREPARACION PRE ESTUDIO
    datos_refuerzos = elementos_criticos(net, analisis_componentes, ranking_contingencias_cb)
    df_cargabilidades_base = unir_duracion(df_duraci, df_cargabilidades_base)
    horas_serie = constantes(df_duraci, datos_estudio['numero_etapas'])
    indice_red_base = indice_cond_n(df_cargabilidades_base, df_duraci, exponente_n_pip, horas_serie)
    
    # RESULTADOS DEL ANALISIS POR PROPUESTA
    analisis_resultados = [[*indice_red_base, 'Red base']]
    comportamiento_cn = []
    
    for i, propuesta in enumerate(datos_refuerzos):
        id_ref, tipo, nombre = propuesta
        print(f'{'='*80}')
        print(f'PROPUESTA: {nombre}')
        print(f'{'='*80}\n')
        
        # CONFIGURACION PARA LAS SIMULACIONES
        nombre_del_estudio = f'Propuesta_{nombre}'
        config_sim = Configuracion_Simulacion(titulo_estudio = nombre_del_estudio)
        config_sim_contingencias = Configuracion_Simulacion_Contingencias(nombre_estudio = nombre_del_estudio,
                                                            exponente_n=exponente_n_pip)
        
        # CREACION DE CARPETAS
        rta_cb_r, rta_top_r, rta_graf_r, rta_ctg, rta_ctg_pip, _ = creacion_carpetas_refuerzos(ruta_refuerzos,
                                                                                    nombre_del_estudio, i)
        
        # CONFIGURACION DE LA RED CON EL REFUERZO
        net_copy = copy.deepcopy(net)
        activar_ref_propuesto (id_ref, tipo, nombre, net_copy)
        reporte_red (net_copy, rta_top_r, 2, 'RED MODIFICADA')
        
        # CONDICION N
        df_cargabilidades_ref, df_flujos_ref = caso_base_completo (net_copy, df_mline, df_mtrafo, df_demanda,
                df_desp_TH, df_desp_ren, Slacks, datos_estudio, df_fechas, nucleos, rta_cb_r, rta_top_r,
                config_sim, False, False)
        analisis_componentes_ref = analisis_caso_base (df_cargabilidades_ref, df_flujos_ref, df_duraci,
                        net_copy, datos_estudio, config_red, rta_cb_r, True, rta_graf_r,
                        nucleos, trafos_limpios)
        del analisis_componentes_ref, df_flujos_ref
        df_cargabilidades_ref = unir_duracion(df_duraci, df_cargabilidades_ref)
        lista_monitoreados = elementos_monitoreo(df_cargabilidades_base, df_cargabilidades_ref, n_elementos,
                                                nombre)
        lista_monitoreados.append(nombre[:-2])
        lista_monitoreados = list(set(lista_monitoreados))
        df_comparativa, fila_res  = comparativa_cn(df_cargabilidades_base, df_cargabilidades_ref, datos_estudio,
                                nombre, horas_serie, lista_monitoreados, rta_cb_r, 1)
        comportamiento_cn.append(fila_res)
        grafica_comparativa_cb(df_comparativa, nombre, rta_cb_r)
        indice_ref = indice_cond_n(df_cargabilidades_ref, df_duraci, exponente_n_pip, horas_serie)
        analisis_resultados.append([*indice_ref, nombre])
        
        # CONTINGENCIAS
        lista_monitoreados.append(nombre)
        indice_severidad_ref = contingencias_refuerzos(config_sim_contingencias, net_copy, df_mline, df_mtrafo, 
                    df_demanda, df_desp_TH, df_desp_ren, Slacks, datos_estudio, df_fechas, rta_ctg_pip,
                    df_duraci, nucleos, lista_monitoreados)
        _, indice_severidad_ref  = analisis_contingencias(indice_severidad_ref, rta_ctg, rta_ctg_pip,
                                                        config_sim_contingencias)
        comparativa_ctg(ranking_contingencias_cb, indice_severidad_ref, nombre, rta_ctg)
        del net_copy
        gc.collect()
    
    # RESULTADOS
    print(f'{'='*80}')
    df_completo = pd.DataFrame(comportamiento_cn).reset_index(drop=True)
    df_completo.to_csv((Path(ruta_refuerzos)/'Resumen_tecnico(Refuerzos_propuestos).csv'), index = False)
    logger.info('Reporte del analisis tecnico para cada refuerzo completado.')
    print(f'{'='*80}')
    return analisis_resultados

# --- LECTURA DE ESTUDIOS PREVIOS ---
def lectura_estudio_previo(ruta_estudio):
    print(f'{'-'*80}')
    # PATRONES DE CARPETAS
    ruta_estudio = Path(ruta_estudio)
    ruta_condicion_n = ruta_estudio / "1. Condicion_n"
    ruta_contingencias = ruta_estudio / "2. Contingencias"
    
    patron_condicion_n = os.path.join(ruta_condicion_n, "LF(cond-n)_*.csv")
    patron_contingencias = os.path.join(ruta_contingencias, "Ranking(cont)_*.csv")
    
    # BUSQUEDA DE ARCHIVOS
    # CONDICION N
    archivos_encontrados = glob.glob(patron_condicion_n)
    if archivos_encontrados:
        ruta_archivo = archivos_encontrados[0]
        df_cargab_cb = pd.read_csv(ruta_archivo)
        logger.info(f"se cargo exitosamente: {os.path.basename(ruta_archivo)}")
    else:
        e = "No se encontro ningun archivo que empiece con: LF(cond-n)_*"
        df_cargab_cb = pd.DataFrame()
        logger.error(e)
    
    # CONTINGENCIAS
    archivos_encontrados = glob.glob(patron_contingencias)
    if archivos_encontrados:
        ruta_archivo = archivos_encontrados[0]
        df_is_cb = pd.read_csv(ruta_archivo)
        logger.info(f"se cargo exitosamente: {os.path.basename(ruta_archivo)}")
    else:
        e = "No se encontro ningun archivo que empiece con: Ranking(cont)_*"
        df_is_cb = pd.DataFrame()
        logger.error(e)
    
    if df_cargab_cb.empty or df_is_cb.empty:
        e = f'La carpeta {ruta_estudio} no cuenta con el diagnostico de la red de transmision.'
        raise ValueError (e)
    print(f'{'='*80}')
    return df_cargab_cb, df_is_cb

# --- LECTURA DE REFUERZOS PROPUESTOS POR EL USUARIO ---
def ruta_refuerzos_usuario(rta_top_prev):
    prints_refuerzos_usuario(rta_top_prev)
    while True:
        ruta = input_log(f'Ingrese la ruta del archivo excel con los refuerzos ({rf'C:\...\Refuerzos.xlsx'}):')
        ruta = ruta.replace('"', '').replace("'", "")
        try:
            ruta_ref = Path(ruta)
            return ruta_ref, ruta_ref.stem
        except:
            e=rf'La direccion ingresada {ruta_ref} no es valida.'
            logger.error(e)
            logger.info('Ingrese nuevamente la ruta.')
            print(f'{'-'*80}')

#--- AGREGAR COMPONENTE ---
def agregar_componente (net_copy, parametros_red, fila):
    nombre = fila ['Nombre_refuerzo']
    id_from = fila['id_bus_from']
    id_to = fila['id_bus_to']
    un_hv = net_copy.bus.at[id_from, 'vn_kv']
    un_lv = net_copy.bus.at[id_to, 'vn_kv']
    if un_hv==un_lv:
        if fila['length_km'] == 0:
            if not net_copy.bus_geodata.empty:
                p1 = net_copy.bus_geodata.loc[id_from]
                p2 = net_copy.bus_geodata.loc[id_to]
                p1 = (p1.y, p1.x)
                p2 = (p2.y, p2.x)
                longitud = haversine(p1, p2, unit=Unit.KILOMETERS)
            else:
                logger.warning(f'La linea {nombre} no tiene longitud y la base de datos no tiene coordenadas de las barras\npara calcular internamente, por lo que se asumira una longitud de 1[km].')
                longitud = 1
        else:
            longitud = fila['length_km']
        r_ohm_km = fila['r[ohm/km]']
        x_ohm_km = fila['x[ohm/km]']
        f_nF_km = (fila['sn[MVAR/km]']/(un_hv**2))*(1E9/(2*np.pi*net_copy.f_hz))
        suceptancia = fila['sn[MVAR/km]']*longitud
        I_mx_kA = fila['P[MW]']/(un_hv*(3**0.5)*parametros_red.Fp)
        n_ternas = fila['N.Ternas']
        pp.create_line_from_parameters(net_copy,
            name=nombre, 
            from_bus=id_from, 
            to_bus=id_to, 
            r_ohm_per_km = r_ohm_km,
            x_ohm_per_km = x_ohm_km,
            c_nf_per_km = f_nF_km,
            max_i_ka = I_mx_kA,
            in_service = True, 
            g_us_per_km = 0,
            type = 'ol',
            df=1,
            length_km = longitud,
            parallel = n_ternas, 
            max_loading_percent=100)
        identificador_costos = [nombre, longitud, float(un_hv), 0, suceptancia, fila['Tipo']]
        return identificador_costos
    else:
        r_ohm = fila['r[ohm/km]']
        x_ohm = fila['x[ohm/km]']
        if pd.isna(r_ohm):
            r_ohm = 0.0
        if pd.isna(x_ohm):
            raise ValueError(f"El transformador {nombre} no tiene x_ohm_km definido.")
        sn_mva = fila['P[MW]'] / parametros_red.Fp
        z_base = un_hv**2 / sn_mva
        z_trafo = ((r_ohm**2) + (x_ohm**2))**0.5
        Ucc_bnom_porcent = (z_trafo / z_base) * 100
        r_bnom_porcent = (r_ohm / z_base) * 100
        pp.create_transformer_from_parameters(net_copy,
            name = nombre,
            hv_bus = fila['id_bus_from'],
            lv_bus = fila['id_bus_to'],
            sn_mva = sn_mva,
            vn_hv_kv = un_hv,
            vn_lv_kv = un_lv,
            vk_percent = Ucc_bnom_porcent,
            vkr_percent = r_bnom_porcent,
            in_service = True,
            pfe_kw = 0,
            i0_percent = 0,
            shift_degree = 0,
            max_loading_percent = 100,
            parallel = 1)
        identificador_costos = [nombre, sn_mva, un_hv, un_lv, 0, fila['Tipo']]
        return identificador_costos

def actualizar_parametros(net_copy, parametros_red, fila, map_lines, map_trafos):
    nombre = fila ['Nombre_refuerzo']
    id_from = fila['id_bus_from']
    id_to = fila['id_bus_to']
    un_hv = net_copy.bus.at[id_from, 'vn_kv']
    un_lv = net_copy.bus.at[id_to, 'vn_kv']
    if un_hv==un_lv:
        idx = map_lines.get(nombre)
        if idx is not None:
            if fila['length_km'] == 0:
                if not net_copy.bus_geodata.empty:
                    p1 = net_copy.bus_geodata.loc[id_from]
                    p2 = net_copy.bus_geodata.loc[id_to]
                    p1 = (p1.y, p1.x)
                    p2 = (p2.y, p2.x)
                    longitud = haversine(p1, p2, unit=Unit.KILOMETERS)
                else:
                    e = f'La linea {nombre} no tiene longitud y la base de datos no tiene coordenadas de las barras\npara calcular internamente, por lo que se asumira una longitud de 1[km].'
                    logger.warning(e)
                    longitud = 1
            else:
                longitud = fila['length_km']
            r_ohm_km = fila['r[ohm/km]']
            x_ohm_km = fila['x[ohm/km]']
            f_nF_km = (fila['sn[MVAR/km]']/(un_hv**2))*(1E9/(2*np.pi*net_copy.f_hz))
            suceptancia = fila['sn[MVAR/km]']*longitud
            I_mx_kA = fila['P[MW]']/(un_hv*(3**0.5)*parametros_red.Fp)
            n_ternas = int(fila['N.Ternas'])
            net_copy.line.at[idx, 'in_service'] = True
            net_copy.line.at[idx, 'r_ohm_per_km'] = r_ohm_km
            net_copy.line.at[idx, 'x_ohm_per_km'] = x_ohm_km
            net_copy.line.at[idx, 'c_nf_per_km'] = f_nF_km
            net_copy.line.at[idx, 'max_i_ka'] = I_mx_kA
            net_copy.line.at[idx, 'parallel'] = n_ternas
            identificador_costos = [nombre, longitud, float(un_hv), 0, suceptancia, fila['Tipo']]
            return identificador_costos
    else:
        idx = map_trafos.get(nombre)
        if idx is not None:
            r_ohm = fila['r[ohm/km]']
            x_ohm = fila['x[ohm/km]']
            if pd.isna(r_ohm):
                r_ohm = 0.0
            if pd.isna(x_ohm):
                raise ValueError(f"El transformador {nombre} no tiene x_ohm_km definido.")
            sn_mva = fila['P[MW]'] / parametros_red.Fp
            z_base = un_hv**2 / sn_mva
            z_trafo = ((r_ohm**2) + (x_ohm**2))**0.5
            Ucc_bnom_porcent = (z_trafo / z_base) * 100
            r_bnom_porcent = (r_ohm / z_base) * 100
            net_copy.trafo.at[idx, 'in_service'] = True
            net_copy.trafo.at[idx, 'vk_percent'] = Ucc_bnom_porcent
            net_copy.trafo.at[idx, 'vkr_percent'] = r_bnom_porcent
            net_copy.trafo.at[idx, 'sn_mva'] = sn_mva
            identificador_costos = [nombre, sn_mva, un_hv, un_lv, 0, fila['Tipo']]
            return identificador_costos

def agregar_barras(net_copy, df_filtrado, nombre_propuesta):
    try:
        df_barras = df_filtrado[df_filtrado['Tipo'] == 'BARRA'].copy()
    except: 
        df_barras = pd.DataFrame()
    
    if df_barras.empty:
        logger.info(f'No se tienen barras en la alternativa {nombre_propuesta}.')
        return df_filtrado
    
    else:
        df_barras['U_kV'] = pd.to_numeric(df_barras['Nombre_refuerzo'].str.extract(r'(\d+)')[0], errors='coerce')
        
        for _, bus in df_barras.iterrows():
            if pd.isna(bus['Longitud']) or pd.isna(bus['Latitud']):
                coordenadas = None
            else:
                coordenadas = (bus['Longitud'], bus['Latitud'])
            id_bus = pp.create_bus(
                net_copy,
                name=bus['Nombre_refuerzo'],
                vn_kv=bus['U_kV'],
                in_service=True,
                type='b', 
                geodata=coordenadas)
            
            df = df_filtrado[df_filtrado['Nombre_refuerzo'] != bus['Nombre_refuerzo']].copy()
            barra, un = str(bus['Nombre_refuerzo']).split('-')
            nom = barra + un
            for idx_fila, fila in df.iterrows():
                nombre_elemento = str(fila['Nombre_refuerzo'])
                if nombre_elemento.startswith(nom):
                    df_filtrado.at[idx_fila, 'id_bus_from'] = id_bus
                if nombre_elemento.endswith(nom):
                    df_filtrado.at[idx_fila, 'id_bus_to'] = id_bus
            logger.info(f"Se creo la barra {bus['Nombre_refuerzo']} con indice {id_bus}.")
            logger.info ('Se actualizaron los "id_bus_from" e "id_bus_to"')
        
        df_sin_barras = df_filtrado[df_filtrado['Tipo'] != 'BARRA'].copy()
        df_verificacion_buses = df_sin_barras[df_sin_barras['En servicio'] == 0].copy()
        if df_verificacion_buses[['id_bus_from', 'id_bus_to']].isna().any().any():
            raise ValueError(
                f"La propuesta {nombre_propuesta} tiene elementos con id_bus_from o id_bus_to vacío. "+
                "Revise nombres de barras nuevas y conexiones en el Excel.")
        return df_sin_barras

def sacar_servicio(df_sin_barras, net_copy):
    try:
        df_fuera_servicio = df_sin_barras[df_sin_barras['En servicio']==1].copy()
    except:
        df_fuera_servicio= pd.DataFrame()
    if not df_fuera_servicio.empty:
        for elemento in (df_fuera_servicio['Nombre_refuerzo'].tolist()):
            idx_line = net_copy.line[net_copy.line['name'] == elemento].index
            if not idx_line.empty:
                net_copy.line.loc[idx_line, 'in_service'] = False
                logger.info(f'Se saco de servicio a la linea {elemento}.')
                continue
            idx_trafo = net_copy.trafo[net_copy.trafo['name'] == elemento].index
            if not idx_trafo.empty:
                net_copy.trafo.loc[idx_trafo, 'in_service'] = False
                logger.info(f'Se saco de servicio al transformador {elemento}.')
                continue
            logger.warning(f"Elemento '{elemento}' no encontrado en la red")
        df_proyectos = df_sin_barras[df_sin_barras['En servicio']!=1].copy()
        return df_proyectos
    else:
        logger.info('No se tienen desconexion de elementos.')
        return df_sin_barras

def elementos_monitoreo_alternativas(df_cargabilidades_base: pd.DataFrame, df_cargabilidades_ref: pd.DataFrame,
                        n_elementos: int, nombre:str, lista_monitoreo: list):
    print(f'{'='*80}')
    print(f'EVALUACION DE LA ALTERNATIVA {nombre} EN CONDICION N.')
    print(f'{'='*80}')
    df_cargab_union = pd.merge(df_cargabilidades_base, df_cargabilidades_ref, how='inner',
                            on=['Etapa', 'Serie', 'Bloque','Componente'],
                            suffixes= ('_base', '_ref'))
    df_cargab_union = df_cargab_union[['Etapa', 'Serie', 'Bloque', 'Componente', 'loading_percent_base', 
                                    'loading_percent_ref']].copy()
    df_cargab_union ['delta_cargab'] = (df_cargab_union['loading_percent_base'] - df_cargab_union['loading_percent_ref']).abs()
    df_maximos = df_cargab_union.groupby('Componente')['delta_cargab'].max().reset_index()
    df_maximos.sort_values(by='delta_cargab', ascending=False, inplace=True)
    df_maximos = df_maximos.head(n_elementos)
    elementos_sensibles = df_maximos['Componente'].to_list()
    if lista_monitoreo:
        logger.info(f'Los {n_elementos} elementos de monitoreo mas la lista de elementos de monitoreo decalarados en'+
                    f'\nexcel para {nombre} son:\n{elementos_sensibles}')
        return list(set(elementos_sensibles))
    logger.info(f'Los {n_elementos} elementos de monitoreo para {nombre} son:\n{elementos_sensibles}')
    return elementos_sensibles

def reporte_tecnico_alternativa(df_comparativa, ranking_contingencias_rb, indice_severidad_ref, nombre_propuesta,
                                ruta_ref):
    print(f'{'-'*80}')
    print(f'REPORTE DEL ANALISIS TECNICO')
    print(f'{'-'*80}')
    
    df_comparativa = pd.merge(df_comparativa, ranking_contingencias_rb, how='left', left_on='Monitoreo', right_on= 'Contingencia')
    df_comparativa.rename(columns={'Ind_Sev':'IS_(Red base)'}, inplace = True)
    df_comparativa = df_comparativa[['Alternativa', 'Monitoreo', 'P_1%(SR)', 'P_5%(SR)', 'P_1%(CR)','P_5%(CR)',
                                    'IS_(Red base)']].copy()
    df_comparativa = pd.merge(df_comparativa, indice_severidad_ref, how='left', left_on='Monitoreo', right_on= 'Contingencia')
    df_comparativa.rename(columns={'Ind_Sev':f'IS_({nombre_propuesta})'}, inplace = True)
    df_comparativa = df_comparativa[['Alternativa', 'Monitoreo', 'P_1%(SR)', 'P_5%(SR)', 'P_1%(CR)','P_5%(CR)',
                                    'IS_(Red base)', f'IS_({nombre_propuesta})']].copy()
    df_comparativa.to_csv((Path(ruta_ref)/f'Reporte_tenico_{nombre_propuesta}.csv'), index=False)
    logger.info(f'Se genero el reporte del analisis tecnico de la alternativa {nombre_propuesta}.')
    print(f'{'-'*80}')
    return df_comparativa

def fecha_restriccion(df_fechas, dias_etapa):
    if dias_etapa ==7:
        etapa_restriccion = 109
    elif dias_etapa ==30:
        etapa_restriccion = 25
    else:
        etapa_restriccion = 9
    try:
        fecha_restriccion = df_fechas.loc[etapa_restriccion, 'Fecha']
    except:
        dias_sum = etapa_restriccion * dias_etapa
        f_inicio = df_fechas.loc[1, 'Fecha']
        fecha_restriccion = f_inicio + timedelta(days = dias_sum)
    finally:
        return fecha_restriccion

def fecha_ingreso(df_cargabilidades_rbase, elemento_monitoreo, df_fechas, dias_etapa, ruta_pips_rb):
    print(f'{'='*80}')
    print(f'DETERMINACION DE FECHAS DE INGRESO PARA LA ALTERNATIVA.')
    print(f'{'='*80}')
    # DF_FECHAS
    df_fechas = df_fechas.copy()
    df_fechas.set_index('Etapa')
    fecha_minima = fecha_restriccion(df_fechas, dias_etapa)
    # PRIMERA FECHA CONDICION N
    df = df_cargabilidades_rbase[df_cargabilidades_rbase['Componente']==elemento_monitoreo].copy()
    df_pivot = df.pivot_table(index=['Etapa', 'Bloque'], columns = 'Serie', values = 'loading_percent')
    df_pivot['Cargab_prom'] = df_pivot.mean(axis=1)
    sobrecarga = df_pivot['Cargab_prom'] >= 100
    if sobrecarga.any():
        indice = sobrecarga.idxmax()
        etapa = indice[0]
        fecha_1 = df_fechas.loc[etapa, 'Fecha']
        if fecha_1 < fecha_minima:
            fecha_1 = fecha_minima
        logger.info('Se identifico la primera fecha de ingreso posible.')
    else:
        logger.info('No se identifico una primera fecha de ingreso posible.')
        logger.info('Por lo que se usara la restriccion de 2 años')
        fecha_1 = fecha_minima
    # SEGUNDA FECHA CONDICION N-1 (CONTINGENCIAS)
    df_pip = pd.read_csv(Path(ruta_pips_rb)/f'PIP_{elemento_monitoreo}.csv')
    df_pivot = df_pip.pivot_table(index=['Etapa', 'Bloque'], columns = 'Serie', values = 'PIp_Total')
    df_pivot['PIP_prom'] = df_pivot.mean(axis=1)
    sobrecarga = df_pivot['PIP_prom'] >= 1.1
    if sobrecarga.any():
        indice = sobrecarga.idxmax()
        etapa = indice[0]
        fecha_2 = df_fechas.loc[etapa, 'Fecha']
        if fecha_2 < fecha_minima:
            fecha_2 = fecha_minima
        logger.info('Se identifico la segunda fecha de ingreso posible.')
    else:
        logger.info('No se identifico una segunda fecha de ingreso posible.')
        logger.info('Por lo que se usara la restriccion de 2 años')
        fecha_2 = fecha_minima
    print(f'{'='*80}')
    return fecha_1, fecha_2

def analisis_cartera(lista_dfs:list, lista_monitoreo: list):
    elementos = [str(elem).strip() for elem in lista_monitoreo]
    datos = {("Elemento monitoreo", ""): elementos}
    _, df_primero = lista_dfs[0]
    df_base = df_primero.copy()
    df_base["Monitoreo"] = df_base["Monitoreo"].astype(str).str.strip()
    columnas_base = [
        "Monitoreo",
        "P_1%(SR)",
        "P_5%(SR)",
        "IS_(Red base)"
    ]
    for col in columnas_base:
        if col not in df_base.columns:
            raise ValueError(f"No se encontró la columna '{col}' en el DataFrame base.")
    df_base = (
        df_base[columnas_base]
        .groupby("Monitoreo", as_index=True)
        .max()
    )
    datos[("Red base", "P1")] = [
        df_base.loc[elem, "P_1%(SR)"] if elem in df_base.index else np.nan
        for elem in elementos
    ]
    datos[("Red base", "P5")] = [
        df_base.loc[elem, "P_5%(SR)"] if elem in df_base.index else np.nan
        for elem in elementos
    ]
    datos[("Red base", "I_Sev")] = [
        df_base.loc[elem, "IS_(Red base)"] if elem in df_base.index else np.nan
        for elem in elementos
    ]
    for nombre_alternativa, df_alt in lista_dfs:
        df = df_alt.copy()
        df["Monitoreo"] = df["Monitoreo"].astype(str).str.strip()
        col_is = f"IS_({nombre_alternativa})"
        columnas_alt = [
            "Monitoreo",
            "P_1%(CR)",
            "P_5%(CR)",
            col_is
        ]
        for col in columnas_alt:
            if col not in df.columns:
                raise ValueError(
                    f"No se encontró la columna '{col}' "
                    f"en el DataFrame de {nombre_alternativa}."
                )
        df = (
            df[columnas_alt]
            .groupby("Monitoreo", as_index=True)
            .max()
        )
        datos[(nombre_alternativa, "P1")] = [
            df.loc[elem, "P_1%(CR)"] if elem in df.index else np.nan
            for elem in elementos
        ]
        datos[(nombre_alternativa, "P5")] = [
            df.loc[elem, "P_5%(CR)"] if elem in df.index else np.nan
            for elem in elementos
        ]
        datos[(nombre_alternativa, "I_Sev")] = [
            df.loc[elem, col_is] if elem in df.index else np.nan
            for elem in elementos
        ]
    df_resumen = pd.DataFrame(datos)
    df_resumen.columns = pd.MultiIndex.from_tuples(df_resumen.columns)
    return df_resumen

def df_resumen_to_excel(df_resumen, ruta_base_cartera):
    ruta_excel = Path(ruta_base_cartera)/'Resumen_tecnico(Alternativas Propuestas).xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Alternativas Propuestas"
    fill_header = PatternFill("solid", fgColor="BFBFBF")
    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080")
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    font_header = Font(bold=True)
    columnas = list(df_resumen.columns)
    ws.cell(row=1, column=1, value="Elemento monitoreo")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    col_excel = 2
    grupo_actual = None
    col_inicio_grupo = 2
    for i, col in enumerate(columnas[1:], start=2):
        grupo, subcol = col
        if grupo_actual is None:
            grupo_actual = grupo
            col_inicio_grupo = i
        elif grupo != grupo_actual:
            ws.cell(row=1, column=col_inicio_grupo, value=grupo_actual)
            ws.merge_cells(
                start_row=1,
                start_column=col_inicio_grupo,
                end_row=1,
                end_column=i - 1
            )
            grupo_actual = grupo
            col_inicio_grupo = i
        ws.cell(row=2, column=i, value=subcol)
    if grupo_actual is not None:
        ws.cell(row=1, column=col_inicio_grupo, value=grupo_actual)
        ws.merge_cells(
            start_row=1,
            start_column=col_inicio_grupo,
            end_row=1,
            end_column=len(columnas)
        )
    for fila_idx, (_, fila) in enumerate(df_resumen.iterrows(), start=3):
        for col_idx, col in enumerate(columnas, start=1):
            ws.cell(row=fila_idx, column=col_idx, value=fila[col])
    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = align_center
            if cell.row <= 2:
                cell.fill = fill_header
                cell.font = font_header
    for row in range(1, max_row + 1):
        ws.cell(row=row, column=1).alignment = align_left
    ws.column_dimensions["A"].width = 25
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "B3"
    wb.save(ruta_excel)

def lista_monitoreo_mayor_1 (lista_isev: list, lista_monitoreo:list, ranking_contingencias_rb: pd.DataFrame)-> tuple[str, int]:
    df = ranking_contingencias_rb[['Contingencia', 'Ind_Sev']].copy()
    df.set_index('Contingencia', inplace = True)
    for elemento in lista_monitoreo:
        if elemento in lista_isev:
            logger.info(f'Se tomara como referencia al elemento {elemento} para el analisis de alternativas.')
            isev_monitoreo_red_base = df.loc[elemento, 'Ind_Sev']
            return elemento, isev_monitoreo_red_base
    elemento = lista_isev[0]
    logger.info(f'No se encontro que ningun elemento de monitoreo exista su indice de contingencia, por lo que se usara: {elemento} para el\n'+
                'analisis de alternativas')
    isev_monitoreo_red_base = df.loc[elemento, 'Ind_Sev']
    return elemento, isev_monitoreo_red_base

def verificar_elem_monitoreo (lista_isev: list, elemento_monitoreo: str, ranking_contingencias_rb: pd.DataFrame)-> tuple[str, int]:
    df = ranking_contingencias_rb[['Contingencia', 'Ind_Sev']].copy()
    df.set_index('Contingencia', inplace = True)
    if elemento_monitoreo in lista_isev:
        logger.info(f'Se tomara como referencia al elemento {elemento_monitoreo} para el analisis de alternativas.')
        isev_monitoreo_red_base = df.loc[elemento_monitoreo, 'Ind_Sev']
        return elemento_monitoreo, isev_monitoreo_red_base
    else:
        logger.info(f'No se encontro que ningun elemento de monitoreo exista su indice de contingencia, por lo que se usara: {elemento_monitoreo} para el\n'+
                'analisis de alternativas')
        isev_monitoreo_red_base = df.loc[lista_isev[0], 'Ind_Sev']
        return lista_isev[0], isev_monitoreo_red_base

def refuerzos_usuario(net, df_refuerzos, df_cargabilidades_rbase, ranking_contingencias_rb, rta_cart, parametros_red,
                    n_elementos, ruta_pips_rb, exponente_n_pip, df_mline, df_mtrafo, df_demanda, df_desp_TH,
                df_desp_ren, Slacks, datos_estudio, df_fechas, df_duraci, config_red, df_costos_ind, df_costos_reactores, nucleos):
    
    # PREPARACION PRE ESTUDIO
    df_cargabilidades_base = unir_duracion(df_duraci, df_cargabilidades_rbase)
    horas_serie = constantes(df_duraci, datos_estudio['numero_etapas'])
    indince_red_base, _ = indice_cond_n(df_cargabilidades_base, df_duraci, exponente_n_pip, horas_serie)
    
    # CARTERA DE PROYECTOS ---
    cartera= set(df_refuerzos['Cartera'].tolist())
    for alternativas in cartera:
        print(f'{'='*60}')
        print(f'Analisis de refuerzos propuestos para la cartera de: {alternativas}')
        print(f'{'='*60}')
        # RUTA CARTERA
        ruta_base_cartera = rta_cart/alternativas
        ruta_base_cartera.mkdir(parents=True, exist_ok=True)
        
        # FILTRAMOS LA CARTERA E IDENTIFICAMOS ALTERNATIVAS
        df_indiv = df_refuerzos[df_refuerzos['Cartera'] == alternativas].copy()
        lista_alternativas = set(df_indiv['Alternativa'].tolist())
        
        # LISTA DE MONITOREO 
        lista_monitoreo = df_indiv['Elementos a monitorear'].dropna().tolist()
        lista_monitoreo = convertir_a_lista(lista_monitoreo)
        lista_isev = ranking_contingencias_rb['Contingencia'].tolist()
        if len (lista_monitoreo)>1:
            elemento_monitoreo, isev_monitoreo_red_base = lista_monitoreo_mayor_1 (lista_isev, lista_monitoreo, ranking_contingencias_rb)
        else:
            elemento_monitoreo, isev_monitoreo_red_base = verificar_elem_monitoreo (lista_isev, lista_monitoreo, ranking_contingencias_rb)
        fila_1 = {'Alternativa':'RB', 'IS(n)':indince_red_base, 'Elemento de monitoreo': elemento_monitoreo,
                        'IS (n-1)':isev_monitoreo_red_base, 'Costo total $': '-', 'Fecha (n)': '-', 'Fecha (n-1)': '-'}
        
        # RESULTADOS POR CARTERA
        presentacion_resultados = []
        df_analisis = []
        presentacion_resultados.append(fila_1)
        
        for num, propuesta in enumerate(sorted(lista_alternativas)):
            nombre_propuesta = str(propuesta).strip()
            net_copy = copy.deepcopy(net)
            costo_proyecto = []
            df_filtrado = df_indiv[df_indiv['Alternativa'] == propuesta].copy()
            
            print(f'{'-'*80}')
            print(f'ALTERNATIVA: {nombre_propuesta}')
            print(f'{'-'*80}')
            
            # BARRAS
            df_sin_barras = agregar_barras(net_copy, df_filtrado, nombre_propuesta)
            
            # RED DE TRANSMISION
            lista_elementos_transmision = net_copy.line['name'].tolist() + net_copy.trafo['name'].tolist()
            map_lines = {name: idx for idx, name in net_copy.line['name'].items()}
            map_trafos = {name: idx for idx, name in net_copy.trafo['name'].items()}
            try:
                df_proyectos = df_sin_barras[df_sin_barras['En servicio']!=1].copy()
            except:
                df_proyectos = df_sin_barras.copy()
            
            for _, fila in df_proyectos.iterrows():
                if fila['Nombre_refuerzo'] in lista_elementos_transmision:
                    id_costos = actualizar_parametros(net_copy, parametros_red, fila, map_lines, map_trafos)
                    costo_proyecto.append(id_costos)
                else:
                    id_costos = agregar_componente(net_copy, parametros_red, fila)
                    costo_proyecto.append(id_costos)
            
            df_proyectos = sacar_servicio(df_sin_barras, net_copy)
            
            # CARPETAS POR ALTERNATIVA
            nombre_carpeta = f'Alternativa_{nombre_propuesta}'
            ruta_caso_base, ruta_reporte_red, ruta_graf, ruta_cont, ruta_pip, ruta_ref = creacion_carpetas_refuerzos(
                ruta_base_cartera, nombre_carpeta, num)
            
            # REPORTE RED
            reporte_red(net_copy, ruta_reporte_red, 2, 'RED MODIFICADA')
            trafos_limpios = trafos_gen(net_copy)
            
            # CONFIGURACION SIMULACION
            config_sim = Configuracion_Simulacion(titulo_estudio=nombre_propuesta)
            config_sim_cont = Configuracion_Simulacion_Contingencias(
                nombre_estudio=nombre_propuesta,
                exponente_n=exponente_n_pip)
            
            # CONDICION N
            df_cargabilidades_ref, df_flujos_ref = caso_base_completo (net_copy, df_mline, df_mtrafo, df_demanda,
                df_desp_TH, df_desp_ren, Slacks, datos_estudio, df_fechas, nucleos, ruta_caso_base, ruta_reporte_red,
                config_sim, False, False)
            analisis_componentes_ref = analisis_caso_base (df_cargabilidades_ref, df_flujos_ref, df_duraci,
                        net_copy, datos_estudio, config_red, ruta_caso_base, True, ruta_graf,
                        nucleos, trafos_limpios)
            del df_flujos_ref, analisis_componentes_ref
            df_cargabilidades_ref = unir_duracion(df_duraci, df_cargabilidades_ref)
            lista_sensibles = elementos_monitoreo_alternativas(df_cargabilidades_base, df_cargabilidades_ref,
                        n_elementos, nombre_propuesta, lista_monitoreo)
            lista_sensibles = list(set(lista_sensibles + lista_monitoreo))
            
            #COMPARATIVA
            df_comparativa  = comparativa_cn(df_cargabilidades_base, df_cargabilidades_ref, datos_estudio,
                                nombre_propuesta, horas_serie, lista_sensibles, ruta_caso_base, 2)
            grafica_comparativa_cb(df_comparativa, nombre_propuesta, ruta_caso_base)
            indice_ref, _ = indice_cond_n(df_cargabilidades_ref, df_duraci, exponente_n_pip, horas_serie)
            
            # CONDICION N-1 (CONTINGENCIAS)
            lista_sensibles = list(set(lista_sensibles) & set(lista_isev))
            indice_severidad_ref = contingencias_refuerzos(config_sim_cont, net_copy, df_mline, df_mtrafo, 
                        df_demanda, df_desp_TH, df_desp_ren, Slacks, datos_estudio, df_fechas, ruta_pip,
                        df_duraci, nucleos, lista_sensibles)
            _, indice_severidad_ref  = analisis_contingencias(indice_severidad_ref, ruta_cont, ruta_pip,
                                                            config_sim_cont)
            comparativa_ctg(ranking_contingencias_rb, indice_severidad_ref, nombre_propuesta, ruta_cont)
            df_aux_contingencias_ref = indice_severidad_ref.copy()
            df_aux_contingencias_ref.set_index('Contingencia', inplace = True)
            isev_monitoreo_alternativa = df_aux_contingencias_ref.loc[elemento_monitoreo, 'Ind_Sev']
            del net_copy, df_aux_contingencias_ref
            
            # REPORTE TECNICO DE LA ALTERNATIVA
            df_comparativa = reporte_tecnico_alternativa(df_comparativa, ranking_contingencias_rb, indice_severidad_ref, nombre_propuesta,
                                ruta_ref)
            df_monitoreo = df_comparativa[df_comparativa['Monitoreo'].isin(lista_monitoreo)].copy()
            df_analisis.append((propuesta, df_monitoreo))
            
            # ANALISIS ECONOMICO
            df_costos = calculo_inversion(costo_proyecto, df_costos_ind, df_costos_reactores, ruta_ref,
                                                nombre_propuesta)
            costo_alternativa = df_costos['Costo_total_$'].sum()
            
            # FECHA DE INGRESO
            fecha_1, fecha_2 = fecha_ingreso(df_cargabilidades_rbase, elemento_monitoreo, df_fechas,
                                            datos_estudio['dias_etapa'], ruta_pips_rb)
            
            fila_x = {'Alternativa':nombre_propuesta, 'IS(n)':indice_ref, 'Elemento de monitoreo': elemento_monitoreo,
                        'IS (n-1)':isev_monitoreo_alternativa, 'Costo total $': costo_alternativa, 'Fecha (n)': fecha_1,
                        'Fecha (n-1)': fecha_2}
            presentacion_resultados.append(fila_x)
        
        # ANALISIS TECNICO (GLOBAL CARTERA)
        df_resumen = analisis_cartera(df_analisis, lista_monitoreo)
        df_resumen_to_excel(df_resumen, ruta_base_cartera)
        
        # RESULTADOS
        df_resultados = pd.DataFrame(presentacion_resultados)
        costo_numerico = pd.to_numeric(df_resultados['Costo total $'], errors='coerce')
        df_resultados['aux_ordenar'] = np.where(df_resultados['Costo total $'] == '-', -np.inf, costo_numerico)
        df_resultados = df_resultados.sort_values(by='aux_ordenar', ascending=True)
        df_resultados = df_resultados.drop(columns=['aux_ordenar'])
        resultados_alternativas_propuestas(ruta_base_cartera, df_resultados, alternativas)