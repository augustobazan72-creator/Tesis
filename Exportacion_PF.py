import logging
import os
import sys
import openpyxl
import pandas as pd
from pathlib import Path
from Configuracion_inicial import input_log
from Menus import menu_digsilent, menu_escenarios, guia__archivo_pareo
from Lector_excels import lector_pareo

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def vinculacion_pf():
    print('='*80)
    print('Vinculacion con PowerFactory (DigSilent)')
    print('='*80)
    while True:
        ruta_usuario = Path(input_log(f'Ingrese la ruta de DIgSIlent(Ej: ({r'C:\...\DIgSILENT\PowerFactory 2024\Python\3.12)'}):\n').strip())
        if ruta_usuario.exists() and ruta_usuario.is_dir():
            try: 
                ruta_dig = ruta_usuario.parent
                os.environ["PATH"] = rf'{str(ruta_dig)}'+';'+os.environ["PATH"]
                sys.path.append(rf'{str(ruta_usuario)}')
                import powerfactory as pf
                app = pf.GetApplication()
                app.Show()
                logger.info('Se vinculo correctamente la API de DigSilent con el programa (Python).')
                print(f'{'-'*80}')
                return(app)
            except:
                logger.warning('Revise que la direccion copiada sea correcta.')
                print(f'{'-'*80}')
        else: 
            logger.warning('La ruta debe ser una la direccion de la carpeta.')
            print(f'{'-'*80}')

def adquisicion_bd_pf(app):
    # Seleccion base de datos
    print('BASE DE DATOS DIGSILENT.')
    logger.info('Del "Data manager" (modo engine) copie la direccion de la base de datos a utilizarse.')
    while True:
        ruta_proyecto = input_log('Ingrese la direccion:\n')
        nombre_usuario, dir_proyecto = ruta_proyecto.split('\\', 1)
        _, nombre_bd = dir_proyecto.split('\\', 1)
        dir_proyecto = dir_proyecto.replace('\\', '\\\\')
        try:
            app.ActivateProject(dir_proyecto)
            # app.Hide()
            logger.info(f'El usuario seleccionado es: {nombre_usuario}')
            logger.info(f'La base de datos (PowerFactory) : {nombre_bd}')
            logger.info('Se activo la base de datos correctamente.')
            print('='*80)
            return dir_proyecto
        except:
            logger.error('La ruta de la base de datos no existe o es incorrecta.')
            print(f'-'*80)

def casos_de_estudio(folder):
    casos = []
    contenido = folder.GetContents()
    for obj in contenido:
        class_name = obj.GetClassName()
        if class_name == 'IntCase':
            casos.append(obj)
        elif class_name == 'IntFolder':
            casos.extend(casos_de_estudio(obj))
    return casos

def casos_estudio(app, dir_proyecto):
    # SELECCION DE CASO DE ESTUDIO
    folder_estudios = app.GetProjectFolder('study', 1)
    if folder_estudios is None:
        logger.error("No se pudo encontrar la carpeta de casos de estudio.")
    else:
        casos = casos_de_estudio(folder_estudios)
        if not casos:
            logger.error("No se encontraron casos de estudio.")
        else:
            print(f'{'='*80}')
            print(f"SELECCION DEL CASO DE ESTUDIO (PF).")
            print(f'{'='*80}')
            logger.info(f"Casos de estudio disponibles: ({len(casos)}):\n")
            for i, caso in enumerate(casos, start=1):
                print(f"{i:>1}. {caso.loc_name}")
            print(f'{'-'*80}')
            while True:
                seleccion = input_log('Ingrese el numero de caso de estudio (a activarse) o "q" para salir: ').strip()
                if seleccion.isdigit() and 1 <= int(seleccion) <= len(casos):
                    caso_elegido = casos[int(seleccion) - 1]
                    caso_elegido.Activate()
                    logger.info(f"Caso de estudio (activado): {caso_elegido.loc_name}")
                    print(f'{'='*80}')
                    return [dir_proyecto.split('\\', -1)[-1], caso_elegido.loc_name]
                elif seleccion.strip().lower():
                    return [None, None]
                else:
                    logger.warning(f"Opcion invalida, ingrese un número entre 1 y {len(casos)}.")

def guardar_valores(ws, lista_PF, lista_SDDP, lista_variaciones):
    # Guardamos lista PF
    for i, elemento in enumerate(lista_SDDP):
        ws[f"A{i+1}"] = elemento
    for i, elemento in enumerate(lista_PF):
        ws[f"B{i+1}"] = elemento
    for i, elemento in enumerate(lista_variaciones):
        ws[f"C{i+1}"] = elemento
    return ws

def elementos_pf(app, ruta, net):
    print('='*80)
    print('EXPORTACION DE ELEMENTOS A EXCEL')
    print('='*80)

    # RED
    print(f'GRIDS')
    folder_red = app.GetProjectFolder('netdat')
    grids = folder_red.GetContents('*.ElmNet')
    for g in grids:
        if g.loc_name.strip() == 'SIN':
            g.outserv = 0
            logger.info(f'Grid activada: {g.loc_name}')
            break
    print(f'{'-'*80}')

    # ACTIVAMOS LOS CASOS DE ESTUDIO 
    folder_estudios = app.GetProjectFolder('study', 1)
    casos = casos_de_estudio(folder_estudios)

    # Diccionarios
    gen_syn_pf = {}
    gen_sta_pf = {}
    loads_pf = {}

    for caso_estudio in casos:
        caso_estudio.Activate()
        nombre_caso = caso_estudio.loc_name
        # GENERADORES SINCRONOS
        generadores_syn_pf = app.GetCalcRelevantObjects('*.ElmSym')
        for gen in generadores_syn_pf:
            if gen.loc_name not in list(gen_syn_pf.keys()):
                gen_syn_pf[gen.loc_name] = [gen, nombre_caso]
        # GENERADORES ESTATICOS
        generadores_sta_pf = app.GetCalcRelevantObjects('*.ElmGenstat')
        for gen in generadores_sta_pf:
            if gen.loc_name not in list(gen_sta_pf.keys()):
                gen_sta_pf[gen.loc_name] = [gen, nombre_caso]
        # DEMANDAS
        cargas_pf = app.GetCalcRelevantObjects('*.ElmLod')
        for carga in cargas_pf:
            if carga.loc_name not in list(loads_pf.keys()):
                loads_pf[carga.loc_name] = [carga, nombre_caso]
    
    # ELEMENTOS - SDDP
    gen_syn_sddp = net.gen['name'].tolist()
    gen_sta_sddp = net.sgen['name'].tolist()
    load_sddp = net.load['name'].tolist()
    
    # PREPARAR LAS LISTAS (PRIMERO SDDP)
    gen_syn_sddp.insert(0, 'SDDP')
    gen_sta_sddp.insert(0, 'SDDP')
    load_sddp.insert(0, 'SDDP')
    # G- SINCRONA (PF)
    nombres_gen_syn = list(gen_syn_pf.keys())
    nombres_gen_syn.insert(0, 'PF')
    variacion_syn = [valor[1] for (_, valor) in gen_syn_pf.items()]
    variacion_syn.insert(0, 'Caso_estudio')
    # G- ESTATICA (PF)
    nombres_gen_sta = list(gen_sta_pf.keys())
    nombres_gen_sta.insert(0, 'PF')
    variacion_sta = [valor[1] for (_, valor) in gen_sta_pf.items()]
    variacion_sta.insert(0, 'Caso_estudio')
    # CARGAS (PF)
    nombre_cargas_pf = list(loads_pf.keys())
    nombre_cargas_pf.insert(0, 'PF')
    variacion_lod = [valor[1] for (_, valor) in loads_pf.items()]
    variacion_lod.insert(0, 'Caso_estudio')
    
    # EXCEL 
    wb = openpyxl.Workbook()
    ws1 = wb.create_sheet('Gen. Syn.', 0)
    ws1 = guardar_valores(ws1, nombres_gen_syn, gen_syn_sddp, variacion_syn)
    ws2 = wb.create_sheet('Gen. Sta.', 1)
    ws2 = guardar_valores(ws2, nombres_gen_sta, gen_sta_sddp, variacion_sta)
    ws3 = wb.create_sheet('Cargas.', 2)
    ws3 = guardar_valores(ws3, nombre_cargas_pf, load_sddp, variacion_lod)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ARCHIVO
    ruta = Path(ruta)/'Pareo_nombres_PF_SDDP.xlsx'
    wb.save(ruta)
    logger.info('Se genero el reporte de elementos correctamente.')
    print('='*80)

def escenarios_disponibles(df, llave):
    print(f'{'='*80}')
    print(f"ESCENARIOS CRITICOS (SDDP)")
    print(f'{'='*80}')
    logger.info(f"Casos de estudio disponibles: ({len(df)}):")
    if llave == 1:
        lista_escenarios = df['Escenarios criticos'].tolist()
    else:
        df['Año'] = df['Año'].astype(str)
        df['escenarios'] = df['Interconexion'] + "_" + df['Lectura'] + "_" + df['Año']
        lista_escenarios = df['escenarios'].tolist()
    for i, caso in enumerate(lista_escenarios, start=1):
        print(f"{i:>1}. {caso}")
    print(f'{'-'*80}')
    while True:
        seleccion = input_log('Ingrese el numero de escenario a importar: ').strip()
        if seleccion.isdigit() and 1 <= int(seleccion) <= len(lista_escenarios):
            escenario_elegido = int(seleccion) - 1
            escenario = df.loc[escenario_elegido] [['Etapa', 'Serie', 'Bloque']]
            print(f'{'='*80}')
            return escenario
        else:
            logger.warning(f"Opcion invalida, ingrese un número entre 1 y {len(lista_escenarios)}.")
            print('-'*80)

def ingresar_escenario():
    while True:
        etapa = input_log('Ingrese la etapa:')
        serie = input_log('Ingrese la serie:')
        bloque = input_log('Ingrese el bloque:')
        escenario = []
        for val in [etapa, serie, bloque]:
            try:
                x = int(val)
                escenario.append(x)
            except:
                e = f'El valor ingresado {x} no es un numero entero.'
                logger.error(e)
        if len(escenario) == 3 :
            return escenario

def base_datos_pf(app):
    # GENERADORES SINCRONOS
    generacion_sincrona = {}
    generadores_syn_pf = app.GetCalcRelevantObjects('*.ElmSym')
    for i in generadores_syn_pf:
        generacion_sincrona[i.loc_name] = i
    # GENERADORES RENOVABLES
    generadores_estaticos = {}
    generadores_sta_pf = app.GetCalcRelevantObjects('*.ElmGenstat')
    for i in generadores_sta_pf:
        if i.GetAttribute('cCategory') != 'Reactive Power Compensation':
            generadores_estaticos[i.loc_name] = i
    # DEMANDAS
    cargas_pf = {}
    cargas = app.GetCalcRelevantObjects('*.ElmLod')
    for i in cargas:
        cargas_pf[i.loc_name] = i
    return generacion_sincrona, generadores_estaticos, cargas_pf

def importar_demanda(escenario, pareo_cargas, app, df_demanda, cargas_pf):
    print('DEMANDA\n')
    # IDENTIFICACMOS LAS CARGAS ACTIVAS EN EL CASO
    cargas = list(cargas_pf.keys())
    df_cargas_activas = pd.DataFrame(cargas, columns=['PF'])
    while True:
        df_cargas_activas = pd.merge(df_cargas_activas, pareo_cargas, on='PF', how='left')
        # REPARTIMOS LA DEMANDA POR CARGAS
        df_demanda = df_demanda.copy()
        df_cargas = df_demanda.loc[[escenario]]
        cargas_sddp = set(pareo_cargas['SDDP'].tolist())
        demandas_pf = []
        for carga in list(cargas_sddp):
            try:
                df = df_cargas_activas[df_cargas_activas['SDDP'] == carga]
                num_demandas = len(df)
                demanda_sddp = float(df_cargas[carga].item())
                demanda_pf = demanda_sddp / num_demandas
                x = [carga, demanda_pf]
                demandas_pf.append(x)
            except:
                continue
        df_demandas_pf = pd.DataFrame(demandas_pf, columns = ['SDDP', 'MW_pf'])
        pareo_cargas = pd.merge(pareo_cargas, df_demandas_pf, on = 'SDDP', how = 'left')
        hay_nulos = df_demandas_pf['SDDP'].isna().any()
        if hay_nulos:
            indices_nulos = df_demandas_pf[df_demandas_pf['SDDP'].isna()].index.tolist()
            logger.warning(f'Se detectaron: {len(indices_nulos)} cargas (PF) sin pareo SDDP.')
            logger.info('Se debe asignar el pareo correpondiente:')
            lista_no_pareados = df_demandas_pf.loc[df_demandas_pf, 'PF'].tolist()
            pareos_usuario = []
            for carga_np in lista_no_pareados:
                while True:
                    pareo_sddp = input_log(f'Ingrese el pareo (SDDP) para {carga_np}:\n')
                    if pareo_sddp.strip().upper() in cargas_sddp:
                        break
                    else:
                        logger.warning('El pareo asignado no esta en la base de datos SDDP.')
                pareo = [carga_np, pareo_sddp]
                pareos_usuario.append(pareo)
            df_nuevos_pareos = pd.DataFrame(pareos_usuario, columns = pareo_cargas.columns)
            pareo_cargas = pd.concat([pareo_cargas, df_nuevos_pareos], ignore_index=True)
            pareo_cargas = pareo_cargas.dropna({'SDDP'})
            logger.info('Se actualizo correctamente el pareo de nombres')
            logger.info('No se olvide actualizar el archivo excel de pareo.')
        else: break
    df_aux = pareo_cargas.copy()
    df_aux.set_index(['PF'], inplace=True)
    for llave, valor in cargas_pf.items():
        p = df_aux.at[llave, 'MW_pf']
        valor.SetAttribute('plini', float(p))
    logger.info('Se cargo exitosamente la demanda SDDP a las cargas en PF')
    print('-'*80)
    return pareo_cargas

def importar_escenarios(pareo_syn, pareo_sta, pareo_cargas, df_p1, df_p2, app, dir_proyecto, df_demanda, df_desp_TH, df_desp_ren):
    opcion = menu_escenarios()
    if opcion == '1':
        escenario = escenarios_disponibles(df_p1, 1)
        gsyn_pf, gsta_pf, cargas_pf = base_datos_pf(app)
        print('='*80)
        print(f'IMPORTACION ESCENARIO.')
        print('='*80)
        pareo_cargas = importar_demanda(escenario, pareo_cargas, app, df_demanda, cargas_pf)
        return False, pareo_cargas
    
    elif opcion == '2':
        escenario = escenarios_disponibles(df_p2, 0)
        return False, pareo_cargas

    elif opcion == '3':
        escenario = ingresar_escenario()
    elif opcion == '4':
        name_bd, name_ce = casos_estudio(app, dir_proyecto)
        if name_bd is None:
            return True, pareo_cargas
        else: return False, pareo_cargas
    else:
        return True, pareo_cargas

def menu_vinculacion_pf(df_p1, df_p2, ruta_escenarios, rta_par, rta_ac, net, df_demanda, df_desp_TH, df_desp_ren):
    # VINCULAMOS CON PF
    app = vinculacion_pf()
    dir_proyecto = adquisicion_bd_pf(app)
    # BUCLE MENU
    while True:
        opcion = menu_digsilent()
        if opcion == '1':
            guia__archivo_pareo()
            pareo_syn, pareo_sta, pareo_cargas = lector_pareo()
            if pareo_syn.empty and pareo_sta.empty and pareo_cargas.empty:
                e = 'Las tablas de pareo estan vacias, Revise el archivo de pareo'
                logger.warning(e)
            else:                
                name_bd, name_ce = casos_estudio(app, dir_proyecto)
                while True:
                    if name_bd is None:
                        break
                    else:
                        salir, pareo_cargas = importar_escenarios(pareo_syn, pareo_sta, pareo_cargas, df_p1, df_p2,
                                                app, dir_proyecto, df_demanda, df_desp_TH, df_desp_ren)
                        if salir:
                            break
                break
        elif opcion =='2':
            elementos_pf(app, rta_par, net)
        else:
            return